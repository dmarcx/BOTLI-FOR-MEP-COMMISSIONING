import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
import fitz
import os

@st.cache_data
def load_data():
    ag = pd.read_csv("Lighting_AboveGround.csv")
    bg = pd.read_csv("Lighting_BelowGround.csv")
    return ag, bg

above_ground, below_ground = load_data()

def get_room_type(room):
    df = above_ground if room.startswith("L") else below_ground
    row = df[df["Room Number"].str.upper().str.strip() == room]
    if row.empty:
        return None, "Room not found"
    room_type = row.iloc[0].get("Type of room", "").strip()
    synonyms = {
        "MEETING ROOM": "×—×“×¨ ×™×©×™×‘×•×ª",
        "MEETINGROOM": "×—×“×¨ ×™×©×™×‘×•×ª",
        "OFFICE": "××©×¨×“",
        "CORRIDOR": "××¡×“×¨×•×Ÿ",
        "ELECTRICAL ROOM": "×—×“×¨ ×—×©××œ",
        "PARKING": "×—× ×™×•×Ÿ",
        "OUTDOOR": "×—×•×¥",
        "RAMP": "×¨××¤×”"
    }
    room_type = synonyms.get(room_type.upper().replace(" ", ""), room_type)
    return room_type, None if room_type else "Room type missing"

def check_documents(room):
    df = above_ground if room.startswith("L") else below_ground
    row = df[df["Room Number"].str.upper().str.strip() == room]
    if row.empty:
        return None
    return row.iloc[0].get("××¡××›×™× ×¡×•×¤×§×•", "").strip() == "×›×Ÿ"

def get_schedule_date(room):
    df = above_ground if room.startswith("L") else below_ground
    row = df[df["Room Number"].str.upper().str.strip() == room]
    if row.empty:
        return None, None, "Room not found"
    planned_date_str = row.iloc[0].get("Commissioning planned date", "").strip()
    if not planned_date_str:
        return None, None, "No planned date found"
    planned_date = datetime.strptime(planned_date_str, "%d-%b-%y").date()
    today = datetime.today().date()
    delta = (today - planned_date).days
    if delta == 0:
        status = "×‘××•×¢×“"
    elif delta > 0:
        status = f"×××•×—×¨×ª ×‘â€{delta} ×™××™×"
    else:
        status = f"××•×§×“××ª ×‘â€{abs(delta)} ×™××™×"
    return planned_date, today, status

def evaluate_lux(room_type, measured_lux):
    lux_table = {
        "×—×“×¨ ×™×©×™×‘×•×ª": 500,
        "××¡×“×¨×•×Ÿ": 200,
        "××©×¨×“": 400,
        "×—×“×¨ ×—×©××œ": 400,
        "×—× ×™×•×Ÿ": 75,
        "×—×•×¥": 30,
        "×¨××¤×”": 300
    }
    required = lux_table.get(room_type)
    if required is None:
        st.warning(f"××–×”×¨×”: ×œ× × ××¦××” ×“×¨×™×©×ª ×œ×•×§×¡ ×¢×‘×•×¨ ×¡×•×’ ×”×—×“×¨ '{room_type}' â€” ×™×™×ª×›×Ÿ ×©×”×¢×¨×š ×©×’×•×™ ××• ×œ× × ×ª××š.")
        return "×¡×•×’ ×—×“×¨ ×œ× ××–×•×”×” â€“ ×“×¨×™×©×•×ª ×”×ª××•×¨×” ××™× ×Ÿ ×™×“×•×¢×•×ª."
    deviation = measured_lux - required
    if measured_lux == 0:
        return "×œ× × ××“×“×” ×¢×•×¦××ª ×”××¨×” â€“ × ×“×¨×© ×œ×”×–×™×Ÿ ×¢×¨×š."
    elif deviation >= 0:
        return "×¨××ª ×”×”××¨×” ×ª×§×™× ×”."
    elif -10 <= deviation < 0:
        return "×¡×˜×™×™×” ×§×œ×” â€“ ×ª×™×¨×©× ×”×¢×¨×” ×œ×™×“×™×¢×ª ×”××ª×›× ×Ÿ."
    else:
        return "×¨××ª ×”×”××¨×” ××™× ×” ×ª×§×™× ×” â€“ × ×“×¨×© ×ª×™×§×•×Ÿ ××• ××™×©×•×¨ ×”××ª×›× ×Ÿ."

def get_power_sources(room):
    if room.startswith("L"):
        file_name = "SLD1-L3-EL-001.pdf"
    elif room.startswith("P"):
        file_name = f"SLD1-P{room[1]}-001.pdf"
    else:
        return []
    if not os.path.exists(file_name):
        return []
    doc = fitz.open(file_name)
    text = "".join([page.get_text() for page in doc])
    return list(set(line.strip() for line in text.splitlines() if "EP-" in line and line.strip().startswith("EP-")))

def get_lighting_fixtures(room):
    df = above_ground if room.startswith("L") else below_ground
    rows = df[df["Room Number"].str.upper().str.strip() == room]
    if rows.empty:
        return ["×œ× × ××¦××• × ×ª×•× ×™×"]
    fixtures = []
    for _, row in rows.iterrows():
        type_c = str(row.get("Type", "")).strip()
        model_d = str(row.get("Model", "")).strip()
        quantity_e = str(row.get("Quantity", "")).strip()
        if type_c or model_d:
            fixtures.append(f"{type_c} {model_d} â€“ ×›××•×ª: {quantity_e}")
    return fixtures if fixtures else ["×œ× ×¦×•×™×™×Ÿ"]

def generate_report(room, room_type, planned, today, status, lux_result, dark_result, sources, participants, remarks):
    wb = load_workbook("×“×•×— ××¡×™×¨×”.xlsx")
    ws = wb.active
    ws["A1"] = f"×—×“×¨ {room} ({room_type})"
    ws["B3"] = str(planned)
    ws["B4"] = str(today)
    ws["C4"] = status
    ws["B6"] = lux_result
    ws["B7"] = dark_result if dark_result else "×œ× × ××“×“"
    ws["B8"] = ", ".join(sources)
    for i, p in enumerate(participants, start=12):
        ws[f"B{i}"] = p
    start_row = 22
    for i, remark in enumerate(remarks):
        ws[f"B{start_row + i}"] = remark
    file_name = f"report_{room}.xlsx"
    wb.save(file_name)
    return file_name

# Streamlit UI
st.title("BOTLI â€“ ×‘×“×™×§×ª ×ª××•×¨×”")
room = st.text_input("×”×–×Ÿ ××¡×¤×¨ ×—×“×¨ (×œ×“×•×’××” L3001):")
if room:
    room = room.upper().strip()
    if len(room) == 5 and room[0] in ["L", "P"] and room[1:].isdigit():
        room_type, error = get_room_type(room)
        if error:
            st.error(error)
        else:
            st.success(f"×¡×•×’ ×—×“×¨: {room_type} | ××¡×¤×¨ ×—×“×¨: {room}")
            if check_documents(room):
                st.info("âœ”ï¸ ×›×œ ×”××¡××›×™× ×”×•×’×©×•.")
            else:
                st.error("âœ–ï¸ × ×“×¨×© ××™×©×•×¨ ×©×›×œ ×”××¡××›×™× ×”×•×’×©×•. ×œ× × ×™×ª×Ÿ ×œ×”××©×™×š.")
                st.stop()

            planned, today, status = get_schedule_date(room)
            st.info(f"ğŸ“… ×”×ª××¨×™×š ×”××ª×•×›× ×Ÿ ×”×•× {planned}, ×”×™×•× {today}, ×”×‘×“×™×§×” {status}.")

            if st.radio("×”×× × ×™×ª×Ÿ ×œ×”×ª×§×“× ×œ×‘×™×¦×•×¢ ×”×‘×“×™×§×” ×‘×¤×•×¢×œ?", ["×›×Ÿ", "×œ×"]) == "×›×Ÿ":
                if st.radio("×”×× ×§×™×™× ××“ ×ª××•×¨×” ×–××™×Ÿ ×œ×‘×™×¦×•×¢ ×”×‘×“×™×§×”?", ["×›×Ÿ", "×œ×"]) == "×›×Ÿ":
                    st.subheader("ğŸ’¡ ×‘×“×™×§×ª ×’×•×¤×™ ×ª××•×¨×”")
                    fixtures = get_lighting_fixtures(room)
                    for fix in fixtures:
                        st.info(fix)
                    match = st.radio("×”×× ××œ×• ×’×•×¤×™ ×”×ª××•×¨×” ×•×”×›××•×ª ×”×§×™×™××™× ×‘×¤×•×¢×œ?", ["×›×Ÿ", "×œ×"])
                    remarks = []
                    if match == "×œ×":
                        actual = st.text_area("×× × ×”×–×Ÿ ××ª ×¡×•×’×™ ×”×’×•×¤×™× ×•×”×›××•×™×•×ª ×›×¤×™ ×©× ××¦××• ×‘×¤×•×¢×œ (×©×•×¨×” ×œ×›×œ ×¤×¨×™×˜)")
                        if actual:
                            remarks += actual.splitlines()
                            st.success("×”×¨×©×™××” ×¢×•×“×›× ×” ×‘×”×¦×œ×—×”.")

                    st.subheader("ğŸ‘¬ ××™ ×”××©×ª×ª×¤×™× ×‘×‘×“×™×§×” ×•××” ×ª×¤×§×™×“×?")
                    participants_text = st.text_area("×× × ×”×–×Ÿ ×¨×©×™××ª ××©×ª×ª×¤×™× ×‘×¤×•×¨××˜ ×©× â€“ ×ª×¤×§×™×“, ×©×•×¨×” ×œ×›×œ ××©×ª×ª×£")
                    participants = [line.strip() for line in participants_text.splitlines() if line.strip()]
                    if not participants:
                        st.warning("× ×“×¨×©×ª ×¨×©×™××ª ××©×ª×ª×¤×™× ×œ×”××©×š.")
                        st.stop()

                    measured = st.number_input("×”×–×Ÿ ××ª ×¨××ª ×”×”××¨×” ×©× ××“×“×” (×‘×œ×•×§×¡):", min_value=0)
                    if measured:
                        lux_result = evaluate_lux(room_type, measured)
                        st.info(lux_result)

                        dark_result = ""
                        darker_area = st.radio("×”×× ×§×™×™××™× ××–×•×¨×™× ×—×©×•×›×™× ×™×•×ª×¨ ×‘×—×“×¨?", ["×œ×", "×›×Ÿ"])
                        if darker_area == "×›×Ÿ":
                            dark_measure = st.number_input("×”×–×Ÿ ××ª ×¨××ª ×”×”××¨×” ×‘××–×•×¨ ×”×—×©×•×š (×‘×œ×•×§×¡):", min_value=0)
                            if dark_measure:
                                dark_result = evaluate_lux(room_type, dark_measure)
                                st.info("×‘××–×•×¨ ×”×—×©×•×š: " + dark_result)

                        sources = get_power_sources(room)
                        st.markdown("### âš¡ ××§×•×¨×•×ª ××¡×¤×§×” ×©× ××¦××• ×‘×ª×•×›× ×™×ª:")
                        if sources:
                            for s in sources:
                                st.write(f"ğŸ”Œ {s}")
                        else:
                            st.write("×œ× × ××¦××• ××§×•×¨×•×ª ××¡×¤×§×”.")

                        signage_match = st.radio("×”×× ×”×©×™×œ×•×˜ ×‘×¤×•×¢×œ ×ª×•×× ×œ×ª×›× ×•×Ÿ?", ["×›×Ÿ", "×œ×"])
                        if signage_match == "×œ×":
                            remarks.append("×©×™×œ×•×˜ ×œ× ×ª×•×× â€“ × ×“×¨×© ×ª×™×§×•×Ÿ ××• ×¢×“×›×•×Ÿ ×ª×›× ×•×Ÿ.")

                        breaker_test = st.radio("×”×× ×”××•×¨ ×›×‘×” ×œ××—×¨ ×”×¤×œ×ª ×”××××ª?", ["×›×Ÿ", "×œ×"])
                        if breaker_test == "×œ×":
                            remarks.append("× ×“×¨×© ×œ×××ª ××ª ×¤×¢×•×œ×ª ×”××××ª â€“ ×”××•×¨ ×œ× ×›×‘×” ×œ××—×¨ ×”×¤×œ×ª×•.")

                        if st.button("ğŸ“„ ×”×¤×§ ×“×•×— ××¡×™×¨×”"):
                            file = generate_report(room, room_type, planned, today, status, lux_result, dark_result, sources, participants, remarks)
                            with open(file, "rb") as f:
                                st.download_button("ğŸ“¥ ×”×•×¨×“ ××ª ×”×“×•×—", data=f, file_name=file)
                    else:
                        st.warning("× ×“×¨×© ×œ×”×–×™×Ÿ ×¢×¨×š ××“×•×“ ×›×“×™ ×œ×”××©×™×š.")
                else:
                    st.warning("× ×“×¨×© ××“ ×ª××•×¨×” ×œ×‘×™×¦×•×¢ ×”×‘×“×™×§×”.")
            else:
                st.warning("×œ× × ×™×ª×Ÿ ×œ×”××©×™×š ×œ×œ× ××™×©×•×¨ ×”×ª×§×“××•×ª.")
    else:
        st.error("××¡×¤×¨ ×”×—×“×¨ ××™× ×• ×ª×§× ×™. ×™×© ×œ×”×–×™×Ÿ ×§×œ×˜ ×‘××‘× ×” L1234 ××• P1234.")
